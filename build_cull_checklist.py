"""
Build a clean printable / clickable manual-cull checklist for the 173 culled mice.
Used to mark each mouse deceased one-by-one in the Transnetyx UI.

Sorted by cage so Brian can batch-process by physical/UI cage.

Output: Cull_Checklist_<DATE>.xlsx
"""

from pathlib import Path
from datetime import date
import json, re
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule

DIR = Path("/Users/briandes/Library/CloudStorage/OneDrive-MichiganMedicine/Desktop/MacDougald Lab/Mouse Colony")
TODAY = date.today()
ARIAL = "Arial"

with open(DIR / "data" / "colony.json") as f:
    colony = json.load(f)

inv = pd.DataFrame(colony["inventory"])
inv["mouse_id"] = inv["mouse_id"].astype(str).str.strip()
inv["genotype"] = inv["genotype"].fillna("").astype(str).str.strip()

# Recompute today's smart-cull plan to get the 199 IDs
priority = {"AdipoGlo": "maintain", "AdipoGlo+": "maintain", "Marrow Glo": "expand",
            "Adipoq-Cre": "winddown", "mTmG": "winddown", "Dendra2": "winddown"}

def is_adipoq_donor(g):
    g = (g or "").lower().replace(" ", ""); return "adipoq-cre" in g and "dendra2<-/->" in g and "mtmg<mtmg" not in g
def is_dendra2_donor(g):
    g = (g or "").lower().replace(" ", ""); return bool(re.search(r"dendra2<\+/[+-]>", g)) and "adipoq-cre" not in g and "mtmg<mtmg" not in g
def is_mtmg_donor(g):
    g = (g or "").lower().replace(" ", ""); return "mtmg<mtmg/mtmg>" in g and "adipoq-cre" not in g and not re.search(r"dendra2<\+/[+-]>", g)

breeders = set()
for b in colony["breedings"]:
    if b["status"] != "Active": continue
    for k in ("mother", "father"):
        v = str(b.get(k, ""))
        s = v.split("|")[0].strip() if "|" in v else v.strip()
        if s: breeders.add(s)
for r in inv.to_dict("records"):
    if r["use"] == "Breeding": breeders.add(r["mouse_id"])

rescue_donors = set()
for target, fn in [("Adipoq-Cre", is_adipoq_donor), ("Dendra2", is_dendra2_donor), ("mTmG", is_mtmg_donor)]:
    if priority[target] != "winddown": continue
    cands = inv[(inv["strain"] != target) & (inv["use"] == "Available") &
                (~inv["mouse_id"].isin(breeders)) & (inv["sex"] != "Unknown")].copy()
    cands = cands[cands["genotype"].apply(fn)]
    for sex in ("Female", "Male"):
        sub = cands[cands["sex"] == sex].dropna(subset=["age_months"]).sort_values("age_months").head(4)
        for mid in sub["mouse_id"]:
            rescue_donors.add(mid)

inv_sorted = inv.sort_values("age_months", na_position="last")
cell_rank = {}
for (s, g, sx), grp in inv_sorted.groupby(["strain", "genotype", "sex"], sort=False):
    for i, mid in enumerate(grp["mouse_id"]):
        cell_rank[mid] = (i, len(grp))
geno_size = inv.groupby(["strain", "genotype"]).size().to_dict()
RESCUABLE = {"Adipoq-Cre", "mTmG", "Dendra2"}

def is_culled(m):
    mid = m["mouse_id"]
    if m["use"] == "Breeding" or mid in breeders: return False
    if m["sex"] == "Unknown": return False
    if mid in rescue_donors: return False
    if geno_size.get((m["strain"], m["genotype"]), 0) == 1 and m["strain"] not in RESCUABLE: return False
    p = priority.get(m["strain"], "maintain")
    if p == "expand": return False
    am = m.get("age_months")
    if p == "winddown": return not (pd.notna(am) and am < 6)
    idx, total = cell_rank.get(mid, (999, 0))
    if idx < min(4, total): return False
    if pd.notna(am) and am < 12: return False
    return True

inv["plan_cull"] = inv.apply(is_culled, axis=1)

UNFOUND = {"0023","100","101","102","130","132","179","207","293","294","295",
           "48","68","85","87","92","141","166","225",
           "00001","00002","00008","00016","0417","202","204"}

killed = inv[(inv["plan_cull"]) & (~inv["mouse_id"].isin(UNFOUND))].copy()
killed = killed.sort_values(["cage_id", "mouse_id"]).reset_index(drop=True)
print(f"Killed mice for checklist: {len(killed)}")

# ---- Workbook ----
wb = Workbook()

# Style
HEADER_FILL = PatternFill("solid", start_color="1E3A8A")
HEADER_FONT = Font(name=ARIAL, size=11, bold=True, color="FFFFFF")
SUBHEADER_FILL = PatternFill("solid", start_color="E0E7FF")
ALT_FILL = PatternFill("solid", start_color="F8FAFC")
DONE_FILL = PatternFill("solid", start_color="DCFCE7")
THIN = Side(border_style="thin", color="CBD5E1")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# ===== Checklist sheet =====
ws = wb.active
ws.title = "Cull Checklist"
ws.sheet_view.showGridLines = False
ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 0
ws.sheet_properties.pageSetUpPr.fitToPage = True
ws.page_margins.left = 0.3; ws.page_margins.right = 0.3
ws.page_margins.top = 0.5; ws.page_margins.bottom = 0.5

# Title block
ws.merge_cells("A1:H1")
ws["A1"] = f"Manual Cull Checklist — Transnetyx Update    ({TODAY.isoformat()})"
ws["A1"].font = Font(name=ARIAL, size=16, bold=True, color="1E3A8A")
ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 28

ws.merge_cells("A2:H2")
ws["A2"] = (f"{len(killed)} mice to mark deceased in Transnetyx Hercules. "
            "Sorted by cage so you can batch through them sequentially. "
            "Tick the 'Done' column as you go — Excel will turn the row green automatically.")
ws["A2"].font = Font(name=ARIAL, size=10, italic=True, color="475569")
ws["A2"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws.row_dimensions[2].height = 30

# Header
hdr = ["✓", "Mouse ID", "Strain", "Sex", "Age (mo)", "Cage", "Genotype", "DOB"]
widths = [5, 13, 14, 6, 9, 14, 38, 12]
for i, h in enumerate(hdr, 1):
    c = ws.cell(row=4, column=i, value=h)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = BORDER
    ws.column_dimensions[get_column_letter(i)].width = widths[i-1]
ws.row_dimensions[4].height = 24
ws.freeze_panes = "A5"
ws.print_title_rows = "1:4"

# Data
for i, m in killed.iterrows():
    r = 5 + i
    ws.cell(row=r, column=1, value="").alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=r, column=2, value=str(m["mouse_id"])).font = Font(name="Consolas", size=10, bold=True)
    ws.cell(row=r, column=3, value=m["strain"])
    ws.cell(row=r, column=4, value="F" if m["sex"]=="Female" else "M" if m["sex"]=="Male" else "U")
    ws.cell(row=r, column=5, value=round(m["age_months"], 1) if pd.notna(m["age_months"]) else "")
    ws.cell(row=r, column=6, value=str(m["cage_id"])).font = Font(name="Consolas", size=10)
    ws.cell(row=r, column=7, value=m["genotype"]).font = Font(name="Consolas", size=9)
    ws.cell(row=r, column=8, value=m["dob"] or "")

    for cc in range(1, 9):
        cell = ws.cell(row=r, column=cc)
        if cc not in (2, 6, 7) or cell.font.name != "Consolas":
            cell.font = Font(name=ARIAL, size=10)
        cell.alignment = Alignment(horizontal="center" if cc in (1,4,5) else "left", vertical="center")
        cell.border = BORDER
        if cc == 1:
            cell.fill = PatternFill("solid", start_color="FFFFFF")
        elif i % 2 == 1:
            cell.fill = ALT_FILL

    ws.row_dimensions[r].height = 18

# Conditional format: when col A has any content, paint the row green
last = 5 + len(killed) - 1
range_str = f"A5:H{last}"
green_rule = FormulaRule(formula=[f'$A5<>""'],
                         fill=DONE_FILL,
                         font=Font(name=ARIAL, size=10, color="166534", strike=True))
ws.conditional_formatting.add(range_str, green_rule)

# ===== Summary sheet =====
ws2 = wb.create_sheet("Summary", 0)  # put first
ws2.sheet_view.showGridLines = False
ws2.column_dimensions["A"].width = 38
ws2.column_dimensions["B"].width = 28

def hdr_(r, t, sz=14, c="1E3A8A"):
    c_ = ws2.cell(row=r, column=1, value=t)
    c_.font = Font(name=ARIAL, size=sz, bold=True, color=c)

def kv(r, k, v):
    ws2.cell(row=r, column=1, value=k).font = Font(name=ARIAL, size=11, bold=True)
    ws2.cell(row=r, column=2, value=v).font = Font(name=ARIAL, size=11)

hdr_(1, f"Manual Cull Checklist — {TODAY.isoformat()}", 16)
ws2.cell(row=2, column=1, value="MacDougald Lab · MacLab Brian Mice").font = \
    Font(name=ARIAL, size=11, italic=True, color="64748B")

hdr_(4, "Counts (formula-driven)", 12)
kv(5, "Total mice to mark deceased", f'=COUNTA(\'Cull Checklist\'!B5:B{last})')
kv(6, "Marked deceased so far (✓ in col A)", f'=COUNTA(\'Cull Checklist\'!A5:A{last})')
kv(7, "Remaining to do", f'=B5-B6')

hdr_(9, "By strain (in this cull batch)", 12)
strains = killed["strain"].value_counts()
for i, (s, n) in enumerate(strains.items()):
    ws2.cell(row=10 + i, column=1, value=s).font = Font(name=ARIAL, size=10)
    ws2.cell(row=10 + i, column=2, value=int(n)).font = Font(name=ARIAL, size=10)

start_row = 10 + len(strains) + 2
hdr_(start_row, "How to use", 12)
notes = [
    "1. Sort: rows are pre-sorted by Cage ID. In the Transnetyx UI, search by cage to",
    "   batch-process multiple mice from the same cage at once — much faster than",
    "   searching by individual mouse ID.",
    "",
    "2. As you mark each mouse deceased in Transnetyx, type any non-blank value",
    "   (e.g., 'x' or '✓') in column A on the Checklist sheet. The row will turn",
    "   green and strike-through automatically, and the counts above will update.",
    "",
    "3. Date of death = " + TODAY.isoformat() + " · Cause = Cull / IACUC-approved euthanasia.",
    "",
    "4. Print landscape (fits on standard letter, header repeats on each page) if",
    "   you prefer paper. Otherwise just keep the file open in a second window.",
    "",
    "5. The dashboard's Smart Cull Plan tab and the Transnetyx Upload Excel both",
    "   contain the same 173 mice — this checklist is for the manual workflow only.",
]
for i, line in enumerate(notes, 1):
    cell = ws2.cell(row=start_row + i, column=1, value=line)
    cell.font = Font(name=ARIAL, size=10, italic=line.startswith("   "))

out = DIR / f"Cull_Checklist_{TODAY.isoformat()}.xlsx"
wb.save(out)
print(f"Saved: {out}")
