"""
Build a Transnetyx-compatible bulk-import Excel that reflects the post-cull state.

Logic:
  1. Load the live inventory (`MacLab - Brian_Mice.xlsx` via colony.json).
  2. Recompute the smart cull plan with the same defaults the dashboard used today
     (Marrow Glo expand, Adipoq-Cre/mTmG/Dendra2 winddown, others maintain;
     keep-floor 4; donor count 4 per sex per orphan strain).
  3. Subtract the 26 mouse IDs Brian could not physically locate.
     → that gives the "killed today" list.
  4. Survivors = all mice − today's kills.
  5. Write a single Excel with the standard fields a Transnetyx Hercules /
     colony bulk import accepts.

Output: MacLab_Brian_Mice_Transnetyx_Upload_<DATE>.xlsx
"""

from pathlib import Path
from datetime import date, datetime
import json, re
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DIR = Path("/Users/briandes/Library/CloudStorage/OneDrive-MichiganMedicine/Desktop/MacDougald Lab/Mouse Colony")
TODAY = date.today()  # 2026-05-08
ARIAL = "Arial"

# ---- Load colony state ----
with open(DIR / "data" / "colony.json") as f:
    colony = json.load(f)

inv = pd.DataFrame(colony["inventory"])
inv["mouse_id"] = inv["mouse_id"].astype(str).str.strip()
inv["genotype"] = inv["genotype"].fillna("").astype(str).str.strip()

# ---- Recompute smart cull plan in Python ----
priority = {"AdipoGlo": "maintain", "AdipoGlo+": "maintain", "Marrow Glo": "expand",
            "Adipoq-Cre": "winddown", "mTmG": "winddown", "Dendra2": "winddown"}
KEEP_FLOOR = 4
DONOR_N = 4

def is_adipoq_donor(g):
    g = (g or "").lower().replace(" ", "")
    return "adipoq-cre" in g and "dendra2<-/->" in g and "mtmg<mtmg" not in g
def is_dendra2_donor(g):
    g = (g or "").lower().replace(" ", "")
    return bool(re.search(r"dendra2<\+/[+-]>", g)) and "adipoq-cre" not in g and "mtmg<mtmg" not in g
def is_mtmg_donor(g):
    g = (g or "").lower().replace(" ", "")
    return "mtmg<mtmg/mtmg>" in g and "adipoq-cre" not in g and not re.search(r"dendra2<\+/[+-]>", g)

# Active breeders
breeders = set()
for b in colony["breedings"]:
    if b["status"] != "Active":
        continue
    for k in ("mother", "father"):
        v = str(b.get(k, ""))
        ids = v.split("|")[0].strip() if "|" in v else v.strip()
        if ids:
            breeders.add(ids)
for r in inv.to_dict("records"):
    if r["use"] == "Breeding":
        breeders.add(r["mouse_id"])

# Rescue donors (top-N youngest per sex per orphan strain)
rescue_donors = {}
for target, fn in [("Adipoq-Cre", is_adipoq_donor), ("Dendra2", is_dendra2_donor), ("mTmG", is_mtmg_donor)]:
    if priority[target] != "winddown":
        continue
    cands = inv[(inv["strain"] != target) & (inv["use"] == "Available") &
                (~inv["mouse_id"].isin(breeders)) & (inv["sex"] != "Unknown")].copy()
    cands = cands[cands["genotype"].apply(fn)]
    for sex in ("Female", "Male"):
        sub = cands[cands["sex"] == sex].dropna(subset=["age_months"]).sort_values("age_months").head(DONOR_N)
        for mid in sub["mouse_id"]:
            rescue_donors[mid] = target

# Per-cell youngest ranking
inv_sorted = inv.sort_values("age_months", na_position="last")
cell_rank = {}
for (s, g, sx), grp in inv_sorted.groupby(["strain", "genotype", "sex"], sort=False):
    for i, mid in enumerate(grp["mouse_id"]):
        cell_rank[mid] = (i, len(grp))

geno_size = inv.groupby(["strain", "genotype"]).size().to_dict()
RESCUABLE = {"Adipoq-Cre", "mTmG", "Dendra2"}

def is_culled_by_plan(m):
    mid = m["mouse_id"]
    if m["use"] == "Breeding" or mid in breeders:
        return False
    if m["sex"] == "Unknown":
        return False
    if mid in rescue_donors:
        return False
    if geno_size.get((m["strain"], m["genotype"]), 0) == 1 and m["strain"] not in RESCUABLE:
        return False
    p = priority.get(m["strain"], "maintain")
    if p == "expand":
        return False
    am = m.get("age_months")
    if p == "winddown":
        return not (pd.notna(am) and am < 6)
    # maintain
    idx, total = cell_rank.get(mid, (999, 0))
    floor = min(KEEP_FLOOR, total)
    if idx < floor:
        return False
    if pd.notna(am) and am < 12:
        return False
    return True

inv["plan_cull"] = inv.apply(is_culled_by_plan, axis=1)

# ---- 26 unfound IDs (uncrossed on Brian's printed list) ----
UNFOUND = ["0023","100","101","102","130","132","179","207","293","294","295",
           "48","68","85","87","92","141","166","225",
           "00001","00002","00008","00016","0417","202","204"]
unfound_set = set(UNFOUND)

# Killed = plan_cull AND not in unfound
killed = inv[(inv["plan_cull"]) & (~inv["mouse_id"].isin(unfound_set))].copy()
unfound_in_plan = inv[inv["mouse_id"].isin(unfound_set)].copy()
survivors = inv[~((inv["plan_cull"]) & (~inv["mouse_id"].isin(unfound_set)))].copy()

# Sanity check
assert len(killed) + len(survivors) == len(inv), "row accounting failed"

print(f"Total inventory pre-cull: {len(inv)}")
print(f"Plan cull list size:      {inv['plan_cull'].sum()}")
print(f"Could not find (unfound): {len(unfound_in_plan)}")
print(f"Actually killed today:    {len(killed)}")
print(f"Survivors (all sources):  {len(survivors)}")

# ---- Build Excel ----
def status_for(m, killed_ids):
    if m["mouse_id"] in killed_ids:
        return "Sacrificed"
    return "Alive"

KILLED_IDS = set(killed["mouse_id"].tolist())

def make_row(m, override_status=None):
    status = override_status or status_for(m, KILLED_IDS)
    dod = TODAY.isoformat() if status == "Sacrificed" else ""
    cause = "Cull / IACUC-approved euthanasia" if status == "Sacrificed" else ""
    notes = ""
    if m["mouse_id"] in unfound_set:
        notes = "VERIFY AT WALKTHROUGH — could not be located on cull day"
    return {
        "AnimalID": m["mouse_id"],
        "Strain": m["strain"],
        "Sex": "F" if m["sex"] == "Female" else "M" if m["sex"] == "Male" else "U",
        "DOB": m["dob"] or "",
        "DOD": dod,
        "Status": status,
        "Cage": m["cage_id"],
        "Genotype": m["genotype"],
        "DamID": "",
        "SireID": "",
        "Notes": notes,
        "CauseOfDeath": cause,
    }

alive_rows = [make_row(m) for _, m in survivors.iterrows()]
killed_rows = [make_row(m) for _, m in killed.iterrows()]
master_rows = [make_row(m) for _, m in inv.iterrows()]

# ---- Workbook ----
wb = Workbook()
HEADER_FILL = PatternFill("solid", start_color="305496")
HEADER_FONT = Font(name=ARIAL, size=10, bold=True, color="FFFFFF")
ALT_FILL = PatternFill("solid", start_color="F2F2F2")
THIN = Side(border_style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

COLS = ["AnimalID","Strain","Sex","DOB","DOD","Status","Cage","Genotype",
        "DamID","SireID","CauseOfDeath","Notes"]
COL_WIDTHS = [12, 14, 5, 12, 12, 12, 14, 36, 10, 10, 28, 38]

def write_data_sheet(ws, rows, title):
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"
    for i, h in enumerate(COLS, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDER
        ws.column_dimensions[get_column_letter(i)].width = COL_WIDTHS[i-1]
    ws.row_dimensions[1].height = 22
    for r_idx, row in enumerate(rows, start=2):
        for c_idx, key in enumerate(COLS, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=row.get(key, ""))
            cell.font = Font(name=ARIAL, size=10)
            cell.alignment = Alignment(vertical="center", wrap_text=False)
            cell.border = BORDER
            if r_idx % 2 == 0:
                cell.fill = ALT_FILL
            if key == "Status":
                if row[key] == "Sacrificed":
                    cell.fill = PatternFill("solid", start_color="FEE2E2")
                    cell.font = Font(name=ARIAL, size=10, color="9C0006", bold=True)
                else:
                    cell.fill = PatternFill("solid", start_color="DCFCE7")
                    cell.font = Font(name=ARIAL, size=10, color="166534", bold=True)
            if key == "Notes" and row[key]:
                cell.font = Font(name=ARIAL, size=10, color="B45309", italic=True)

# Sheet 1: Alive (post-cull)
ws_alive = wb.active
ws_alive.title = "Alive"
write_data_sheet(ws_alive, alive_rows, "Alive")

# Sheet 2: Sacrificed today
ws_sac = wb.create_sheet("Sacrificed")
write_data_sheet(ws_sac, killed_rows, "Sacrificed")

# Sheet 3: Master (all mice with Status column)
ws_master = wb.create_sheet("Master")
write_data_sheet(ws_master, master_rows, "Master")

# Sheet 4: Verify_Unfound (the 26 to chase down)
ws_verify = wb.create_sheet("Verify_Unfound")
verify_rows = [make_row(m) for _, m in unfound_in_plan.iterrows()]
write_data_sheet(ws_verify, verify_rows, "Verify_Unfound")

# Sheet 5: Summary / Instructions
ws_sum = wb.create_sheet("Summary", 0)  # put first
ws_sum.sheet_view.showGridLines = False
ws_sum.column_dimensions["A"].width = 38
ws_sum.column_dimensions["B"].width = 24

def hdr(row, text, size=14):
    c = ws_sum.cell(row=row, column=1, value=text)
    c.font = Font(name=ARIAL, size=size, bold=True, color="305496")

def kv(row, label, value):
    a = ws_sum.cell(row=row, column=1, value=label)
    a.font = Font(name=ARIAL, size=10, bold=True)
    b = ws_sum.cell(row=row, column=2, value=value)
    b.font = Font(name=ARIAL, size=10)

def text(row, msg, italic=False, color="000000"):
    c = ws_sum.cell(row=row, column=1, value=msg)
    c.font = Font(name=ARIAL, size=10, italic=italic, color=color)

hdr(1, "MacLab — Brian's Mice — Transnetyx Bulk Import", size=14)
text(2, f"Generated {TODAY.isoformat()} for upload to Transnetyx Hercules / Colony Manager", italic=True, color="555555")

hdr(4, "Counts (formula-driven from each sheet)", size=11)
kv(5, "Total mice in records", f"=COUNTA(Master!A2:A{len(master_rows)+1})")
kv(6, "Currently alive (incl. unfound to verify)", f"=COUNTA(Alive!A2:A{len(alive_rows)+1})")
kv(7, "Sacrificed on cull day", f"=COUNTA(Sacrificed!A2:A{len(killed_rows)+1})")
kv(8, "Unfound during cull (need walkthrough)", f"=COUNTA(Verify_Unfound!A2:A{len(verify_rows)+1})")

hdr(10, "How to upload to Transnetyx", size=11)
notes_lines = [
    "1.  Confirm with your Transnetyx rep which template version they accept",
    "    for bulk import. Most Hercules/Colony Manager subscriptions accept",
    "    a tab-delimited or .xlsx file with these standard columns:",
    "       AnimalID, Strain, Sex, DOB, DOD, Status, Cage, Genotype,",
    "       DamID, SireID, CauseOfDeath, Notes",
    "    These columns are present here on every data sheet.",
    "",
    "2.  Two upload approaches depending on what the rep prefers:",
    "    A) Single 'Master' sheet — every mouse with its current Status.",
    "       Use this if Transnetyx accepts an upsert/update import.",
    "    B) Two separate uploads — 'Alive' (the new active roster) plus",
    "       'Sacrificed' (mice to mark deceased). Use this if Transnetyx",
    "       requires separate live and decedent imports.",
    "",
    "3.  If the rep returns a column-mismatch error: rename the headers in",
    "    the Master sheet to match their template exactly (capitalization",
    "    and spelling matter), then re-upload.",
    "",
    "4.  The 'Verify_Unfound' sheet lists mice that could not be located",
    "    physically on the cull day. They are still marked Status=Alive",
    "    but flagged in the Notes column. After tomorrow's walkthrough,",
    "    update Status=Sacrificed (with DOD) for any confirmed dead, and",
    "    leave the rest as Alive.",
    "",
    "5.  DamID and SireID columns are blank — the colony export does not",
    "    include parent links. If Transnetyx requires them, they can be",
    "    backfilled from the Breedings.xlsx export (separately).",
    "",
    "6.  After upload, refresh the dashboard:",
    "       python build_dashboard_data.py && python build_dashboard_html.py",
    "    then commit + push. The live URL will reflect the new state.",
]
for i, line in enumerate(notes_lines, start=11):
    text(i, line, italic=line.startswith("    "))

# Strain breakdown table
breakdown_start = 11 + len(notes_lines) + 2
hdr(breakdown_start, "Strain breakdown after cull", size=11)
sb_row = breakdown_start + 1
ws_sum.cell(row=sb_row, column=1, value="Strain").font = Font(name=ARIAL, size=10, bold=True)
ws_sum.cell(row=sb_row, column=2, value="Alive").font = Font(name=ARIAL, size=10, bold=True)
ws_sum.cell(row=sb_row, column=3, value="Sacrificed").font = Font(name=ARIAL, size=10, bold=True)
ws_sum.column_dimensions["C"].width = 14
sb_row += 1
strain_order = ["AdipoGlo","AdipoGlo+","Dendra2","Marrow Glo","mTmG","Adipoq-Cre"]
for s in strain_order:
    ws_sum.cell(row=sb_row, column=1, value=s).font = Font(name=ARIAL, size=10)
    ws_sum.cell(row=sb_row, column=2,
                value=f'=COUNTIF(Alive!B2:B{len(alive_rows)+1},"{s}")').font = Font(name=ARIAL, size=10)
    ws_sum.cell(row=sb_row, column=3,
                value=f'=COUNTIF(Sacrificed!B2:B{len(killed_rows)+1},"{s}")').font = Font(name=ARIAL, size=10)
    sb_row += 1

# Save
out = DIR / f"MacLab_Brian_Mice_Transnetyx_Upload_{TODAY.isoformat()}.xlsx"
wb.save(out)
print(f"\nSaved: {out}")
