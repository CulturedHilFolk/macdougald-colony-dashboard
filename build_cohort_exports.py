"""
Build cohort assignment Excel exports for:
  1. LPS Ex 2 (Brian's in-house experiment) — option #1 WT definition (any no-Cre)
  2. Annemarie Lang collaborator cohort — option #2 WT definition (strict Dendra2<-/->)

Also writes data/saved_cohorts.json so the dashboard can surface them later.

Standing conventions:
  - mTmG cassette filtering: Lang cohort excludes any mouse with functional mTmG
    cassette + Adipoq-Cre (avoids membrane-mGFP confound in Dendra2 green channel).
  - Littermate-matched WT:EXP pairs preserved where possible.
  - Sex-balanced 1:1 in both cohorts.
  - Age balance attempted; flagged where constrained.

Output files:
  - Mouse_Cohort_LPS_Ex2_2026-06-05.xlsx
  - Mouse_Cohort_Lang_TibialFracture_2026-06-05.xlsx
  - data/saved_cohorts.json
"""
from __future__ import annotations
import json
from pathlib import Path
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE = Path(__file__).parent
COLONY = json.load(open(HERE / "data" / "colony.json"))
INV = {m["mouse_id"]: m for m in COLONY["inventory"]}

TODAY = "2026-06-05"

# ---------- Cohort definitions ----------

LPS_EXP = [
    # young (≤8 mo)
    {"id": "533", "age_class": "young", "litter_pair": "—",      "treatment_slot": "PBS"},
    {"id": "477", "age_class": "young", "litter_pair": "—",      "treatment_slot": "PBS"},
    {"id": "479", "age_class": "young", "litter_pair": "—",      "treatment_slot": "LPS"},
    {"id": "531", "age_class": "young", "litter_pair": "—",      "treatment_slot": "PBS"},
    {"id": "472", "age_class": "young", "litter_pair": "—",      "treatment_slot": "LPS"},
    # old (≥12 mo)
    {"id": "329", "age_class": "old",   "litter_pair": "★ 330 WT", "treatment_slot": "LPS"},
    {"id": "445", "age_class": "old",   "litter_pair": "★ 446 WT", "treatment_slot": "LPS"},
    {"id": "166", "age_class": "old",   "litter_pair": "★ 165 WT", "treatment_slot": "PBS"},
    {"id": "309", "age_class": "old",   "litter_pair": "—",        "treatment_slot": "PBS"},
    {"id": "382", "age_class": "old",   "litter_pair": "★ 381 WT", "treatment_slot": "LPS"},
    {"id": "92",  "age_class": "old",   "litter_pair": "—",        "treatment_slot": "LPS"},
    {"id": "374", "age_class": "old",   "litter_pair": "—",        "treatment_slot": "PBS"},
]
LPS_WT = [  # gate controls
    {"id": "446", "age_class": "old",   "litter_pair": "★ 445 EXP", "treatment_slot": "PBS"},
    {"id": "330", "age_class": "old",   "litter_pair": "★ 329 EXP", "treatment_slot": "LPS"},
    {"id": "165", "age_class": "old",   "litter_pair": "★ 166 EXP", "treatment_slot": "PBS"},
    {"id": "381", "age_class": "old",   "litter_pair": "★ 382 EXP", "treatment_slot": "LPS"},
    {"id": "518", "age_class": "young", "litter_pair": "—",         "treatment_slot": "PBS"},
    {"id": "529", "age_class": "young", "litter_pair": "—",         "treatment_slot": "LPS"},
]

LANG_EXP = [
    # young clean (no functional mTmG)
    {"id": "525", "age_class": "young", "litter_pair": "—"},
    {"id": "519", "age_class": "young", "litter_pair": "—"},
    {"id": "484", "age_class": "young", "litter_pair": "—"},
    {"id": "476", "age_class": "young", "litter_pair": "—"},
    # old clean
    {"id": "335", "age_class": "old",   "litter_pair": "★ 320 WT (cross-sex littermate)"},
    {"id": "337", "age_class": "old",   "litter_pair": "★ 322 WT (cross-sex littermate)"},
    {"id": "351", "age_class": "old",   "litter_pair": "trio with 352/353"},
    {"id": "352", "age_class": "old",   "litter_pair": "trio with 351/353"},
    {"id": "353", "age_class": "old",   "litter_pair": "trio with 351/352"},
    {"id": "347", "age_class": "old",   "litter_pair": "—"},
]
LANG_WT = [  # strict no-Cre + Dendra2<-/->
    {"id": "494", "age_class": "young", "litter_pair": "—"},
    {"id": "333", "age_class": "old",   "litter_pair": "—"},
    {"id": "341", "age_class": "old",   "litter_pair": "—"},
    {"id": "378", "age_class": "old",   "litter_pair": "—"},
    {"id": "119", "age_class": "old",   "litter_pair": "—"},
    {"id": "499", "age_class": "young", "litter_pair": "—"},
    {"id": "407", "age_class": "mid",   "litter_pair": "—"},
    {"id": "320", "age_class": "old",   "litter_pair": "★ 335 EXP (cross-sex littermate)"},
    {"id": "322", "age_class": "old",   "litter_pair": "★ 337 EXP (cross-sex littermate)"},
    {"id": "317", "age_class": "old",   "litter_pair": "—"},
]

# ---------- Helpers ----------

def enrich(entry, group_label):
    m = INV[entry["id"]]
    out = {
        "mouse_id": entry["id"],
        "group": group_label,
        "age_class": entry["age_class"],
        "sex": m.get("sex"),
        "age_months": round(m.get("age_months", 0), 1),
        "dob": m.get("dob"),
        "cage_id": m.get("cage_id"),
        "strain": m.get("strain"),
        "genotype": m.get("genotype"),
        "use": m.get("use"),
        "litter_pair": entry.get("litter_pair", "—"),
    }
    if "treatment_slot" in entry:
        out["treatment_slot"] = entry["treatment_slot"]
    return out

LPS_EXP_E = [enrich(e, "Adipoq-Cre+; Dendra2+ (EXP)") for e in LPS_EXP]
LPS_WT_E  = [enrich(e, "WT gate ctrl (no Adipoq-Cre)") for e in LPS_WT]
LANG_EXP_E = [enrich(e, "Adipoq-Cre+; Dendra2+ (EXP)") for e in LANG_EXP]
LANG_WT_E  = [enrich(e, "Strict WT (no Cre; Dendra2-/-)") for e in LANG_WT]

# ---------- Styling ----------

H_FILL = PatternFill("solid", fgColor="00274C")  # UM blue
H_FONT = Font(bold=True, color="FFFFFF", size=11)
SUB_FILL = PatternFill("solid", fgColor="FFCB05")  # maize
EXP_FILL = PatternFill("solid", fgColor="EAF2FB")
WT_FILL  = PatternFill("solid", fgColor="FFF7E0")
MATCH_FONT = Font(bold=True, color="00274C")
THIN = Side(border_style="thin", color="CFD6E0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

COL_WIDTHS = {
    "mouse_id": 11, "group": 30, "treatment_slot": 12, "age_class": 10,
    "sex": 6, "age_months": 8, "dob": 12, "cage_id": 14,
    "strain": 11, "genotype": 55, "use": 11, "litter_pair": 28,
}

def style_header_row(ws, row_idx, headers):
    for j, h in enumerate(headers, 1):
        c = ws.cell(row=row_idx, column=j, value=h.replace("_", " ").title())
        c.fill = H_FILL
        c.font = H_FONT
        c.alignment = CENTER
        c.border = BORDER

def write_table(ws, start_row, headers, rows, exp_label_substr="EXP"):
    style_header_row(ws, start_row, headers)
    for i, r in enumerate(rows, start=start_row + 1):
        is_exp = exp_label_substr in r.get("group", "")
        for j, h in enumerate(headers, 1):
            v = r.get(h, "")
            c = ws.cell(row=i, column=j, value=v)
            c.alignment = WRAP if h in ("genotype", "group", "litter_pair") else CENTER
            c.border = BORDER
            c.fill = EXP_FILL if is_exp else WT_FILL
            if h == "litter_pair" and isinstance(v, str) and "★" in v:
                c.font = MATCH_FONT
    # column widths
    for j, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(j)].width = COL_WIDTHS.get(h, 14)
    return start_row + 1 + len(rows)

def add_summary_block(ws, row, title, rows):
    ws.cell(row=row, column=1, value=title).font = Font(bold=True, size=12, color="00274C")
    row += 1
    F = sum(1 for r in rows if r["sex"] == "Female")
    M = sum(1 for r in rows if r["sex"] == "Male")
    y = sum(1 for r in rows if r["age_class"] == "young")
    mid = sum(1 for r in rows if r["age_class"] == "mid")
    o = sum(1 for r in rows if r["age_class"] == "old")
    matched = sum(1 for r in rows if "★" in str(r.get("litter_pair", "")))
    cells = [
        ("Total", len(rows)),
        ("Female", F),
        ("Male", M),
        ("Young (≤8 mo)", y),
        ("Mid (8–12 mo)", mid),
        ("Old (≥12 mo)", o),
        ("Littermate-matched", matched),
    ]
    for j, (lab, val) in enumerate(cells, 1):
        ws.cell(row=row, column=j, value=lab).font = Font(bold=True, size=9, color="5C6470")
        ws.cell(row=row, column=j, value=lab).alignment = CENTER
        ws.cell(row=row + 1, column=j, value=val).font = Font(bold=True, size=14, color="00274C")
        ws.cell(row=row + 1, column=j, value=val).alignment = CENTER
        ws.column_dimensions[get_column_letter(j)].width = max(ws.column_dimensions[get_column_letter(j)].width or 12, 14)
    return row + 3

def write_notes_sheet(ws, lines):
    ws.column_dimensions["A"].width = 110
    r = 1
    for line in lines:
        c = ws.cell(row=r, column=1, value=line)
        if line.startswith("### "):
            c.font = Font(bold=True, size=12, color="00274C")
            c.value = line[4:]
        elif line.startswith("## "):
            c.font = Font(bold=True, size=14, color="00274C")
            c.value = line[3:]
        elif line.startswith("# "):
            c.font = Font(bold=True, size=16, color="00274C")
            c.value = line[2:]
        else:
            c.font = Font(size=10)
            c.alignment = Alignment(wrap_text=True, vertical="top")
        r += 1

# ---------- Build LPS Ex 2 workbook ----------

def build_lps_workbook(out_path: Path):
    wb = Workbook()

    # --- Cohort sheet
    ws = wb.active
    ws.title = "Cohort"
    ws.cell(row=1, column=1, value=f"LPS Ex 2 cohort — assembled {TODAY}").font = Font(bold=True, size=14, color="00274C")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=11)

    ws.cell(row=2, column=1, value=(
        "Design: 12 EXP (Adipoq-Cre+; Dendra2+) split PBS×6 / LPS×6  +  6 WT gate controls (PBS×3 / LPS×3).  "
        "WT definition: option #1 (any mouse without Adipoq-Cre).  "
        "All EXP are mTmG-cassette-clean (no functional mG green-channel background)."
    )).alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=11)
    ws.row_dimensions[2].height = 38

    headers = ["mouse_id", "group", "treatment_slot", "age_class", "sex", "age_months",
               "dob", "cage_id", "strain", "genotype", "litter_pair"]

    row = 4
    row = write_table(ws, row, headers, LPS_EXP_E + LPS_WT_E)
    row += 2
    row = add_summary_block(ws, row, "Summary — EXP", LPS_EXP_E)
    row = add_summary_block(ws, row, "Summary — WT gate ctrls", LPS_WT_E)

    # freeze header
    ws.freeze_panes = "A5"

    # --- Notes sheet
    notes = wb.create_sheet("Notes")
    write_notes_sheet(notes, [
        "# LPS Ex 2 — cohort selection notes",
        "",
        "## Experimental design (from Brian's deck)",
        "12 EXP (Adipoq-Cre+; Dendra2+) split into PBS (n=6) and LPS (n=6) arms",
        "6 WT gate controls split into PBS (n=3) and LPS (n=3)",
        "Readout: peripheral blood Dendra2+ signal in circulating cells after in-vivo LPS challenge",
        "Reference: AdGlo LPS/clodronate flow series; transfer ~13–18× induced vs PBS in pilot",
        "",
        "## WT definition used",
        "Option #1: any mouse without Adipoq-Cre (broad pool, 108 candidates colony-wide)",
        "Rationale: WT mice here serve only as flow gate controls (no Adipoq-Cre activity, no",
        "Dendra2 expression possible even when cassette present), so Dendra2 cassette status",
        "does not need to be strict negative.",
        "",
        "## EXP cohort composition",
        "Sex: 6F + 6M",
        "Age: 5 young (≤8 mo) + 7 old (≥12 mo). Slight skew toward old because young-Cre+ M",
        "      pool is small after excluding mTmG-cassette carriers (kept for clean readout).",
        "",
        "## WT cohort composition",
        "Sex: 3F + 3M",
        "Age: 2 young + 4 old. Older skew preserves littermate matching (4 of 6 WT are EXP",
        "      littermates, see star-marked rows).",
        "",
        "## Littermate-matched pairs (★)",
        "329 ♀ EXP ↔ 330 ♀ WT  (cage 1003121676)",
        "445 ♀ EXP ↔ 446 ♀ WT  (cage 1003172013)",
        "166 ♀ EXP ↔ 165 ♀ WT  (cage 1003059365)",
        "382 ♂ EXP ↔ 381 ♂ WT  (cage 1003121780)",
        "",
        "## Treatment-slot assignments (PBS vs LPS) — review before injection",
        "Sex- and age-balanced across both arms. Littermate pairs split across PBS/LPS so each",
        "treatment arm sees the same maternal-environment representation.",
        "",
        "## mTmG cassette filter",
        "All EXP mice are mTmG<WT/WT> or no mTmG cassette → no Cre-induced membrane GFP.",
        "This keeps the green channel clean for Dendra2 detection. (Excluded mouse 473 from",
        "the original draft because it is mTmG<mTmG/mTmG>; Adipoq-Cre+.)",
        "",
        "## Cages that will empty after this pull",
        "1003121676 (329, 330) — both pulled",
        "1003172013 (445, 446) — both pulled",
        "1003059365 (166, 165) — both pulled",
        "1003121780 (382, 381) — both pulled (verify other occupants first)",
        "1003054428 (309) — single-housed, cage empties",
        "1003089166 (374) — single-housed, cage empties",
        "1002947895 (92)  — single-housed, cage empties",
        "",
        "## Replaceable backups",
        "92 ♂ 18.5 mo → 374 ♂ 12.5 mo (if 92 in poor body condition)",
        "374 ♂        → 347 ♂ 12.6 mo",
        "518 / 529 ♂ young WT → many young no-Cre M available; lowest priority to swap",
    ])

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"[lps] wrote {out_path}")

# ---------- Build Lang workbook ----------

def build_lang_workbook(out_path: Path):
    wb = Workbook()

    ws = wb.active
    ws.title = "Cohort"
    ws.cell(row=1, column=1, value=f"Annemarie Lang collaborator cohort — assembled {TODAY}").font = Font(bold=True, size=14, color="00274C")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=11)

    ws.cell(row=2, column=1, value=(
        "Design: 10 EXP (Adipoq-Cre+; Dendra2+, mTmG-clean) + 10 strict-WT (no Cre AND Dendra2<-/->). "
        "Sex-balanced 5F:5M each arm. Purpose: tibial fracture model with detection of adipocyte-derived "
        "mito-Dendra2 in cells recruited to the injury site (hypothesized early erythroid progenitors + adipocytes)."
    )).alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=11)
    ws.row_dimensions[2].height = 42

    headers = ["mouse_id", "group", "age_class", "sex", "age_months",
               "dob", "cage_id", "strain", "genotype", "litter_pair"]

    row = 4
    row = write_table(ws, row, headers, LANG_EXP_E + LANG_WT_E)
    row += 2
    row = add_summary_block(ws, row, "Summary — EXP", LANG_EXP_E)
    row = add_summary_block(ws, row, "Summary — strict WT controls", LANG_WT_E)
    ws.freeze_panes = "A5"

    # --- Recipient briefing (for Annemarie)
    brief = wb.create_sheet("Briefing for Annemarie")
    write_notes_sheet(brief, [
        "# Cohort briefing — Adipoq-Cre × Dendra2 mice for tibial fracture / mito-transfer assay",
        "",
        "From: Brian Desrosiers (MacDougald Lab, University of Michigan)",
        f"Cohort assembled: {TODAY}",
        "Total mice: 20  (10 EXP + 10 WT controls)",
        "",
        "## Genetic background",
        "EXP arm:  Adipoq-Cre+; CAG-LSL-Dendra2+  (adipocyte-restricted mito-Dendra2 expression)",
        "WT arm:   no Adipoq-Cre  AND  Dendra2<-/-> (no cassette present)",
        "",
        "All EXP mice are mTmG-cassette-clean (mTmG<WT/WT> or no mTmG cassette) — this is intentional.",
        "Mice carrying functional mTmG cassette × Adipoq-Cre express membrane GFP (mG) in adipocytes,",
        "which sits in the same emission window as unconverted mito-Dendra2 (~507 nm). Selecting clean",
        "mice removes that confound from your fracture-site Dendra2 readout.",
        "",
        "## Strain notes",
        "Most mice are on the AdipoGlo / AdipoGlo+ background (mixed B6/129). Cage card and Transnetyx",
        "records confirm Adipoq-Cre and Dendra2 status for every mouse listed. mTmG genotype is also",
        "reported in the 'genotype' column for full transparency.",
        "",
        "## Sex and age balance",
        "EXP: 5F + 5M  ·  4 young (≤8 mo) + 6 old (≥12 mo)",
        "WT:  5F + 5M  ·  3 young (≤8 mo) + 6 old (≥12 mo) + 1 mid (11 mo)",
        "Age skew toward older mice on the WT side reflects colony availability — strict-Dendra2<-/->",
        "young mice are scarce. If young controls are essential, please let Brian know and he can plan",
        "a future breeding round to expand the young strict-WT pool.",
        "",
        "## Littermate-matched pairs",
        "335 ♀ EXP ↔ 320 ♂ WT  (cross-sex littermates from cage 1003059439/40 family)",
        "337 ♀ EXP ↔ 322 ♂ WT  (cross-sex littermates from cage 1003059439/40 family)",
        "351, 352, 353 ♂ EXP — sibling trio from one litter (cage 1003121755), useful as a within-cohort",
        "                       biological-replicate set.",
        "",
        "## Suggested handling",
        "Ear-tag IDs are tattooed and Transnetyx-verified. Shipping crate or hand-off arrangements TBD.",
        "If you want any of the older WT controls replaced with younger animals from a future litter,",
        "we can set that up — current pool is age-limited by strict-Dendra2<-/-> requirement.",
        "",
        "## Contact",
        "Brian Desrosiers  ·  briandesrosiers6@gmail.com  ·  MacDougald Lab, UM Internal Medicine",
    ])

    # --- Internal notes (Brian's records)
    notes = wb.create_sheet("Internal notes")
    write_notes_sheet(notes, [
        "# Lang collaborator cohort — internal selection notes",
        "",
        "## WT definition used",
        "Option #2 (strict): no Adipoq-Cre AND Dendra2<-/->. Colony pool = 25 mice.",
        "",
        "## Why strict (vs. option #1 used for LPS Ex 2)",
        "Lang's assay reads Dendra2 fluorescence in cells *at the fracture site*. Any Dendra2 cassette",
        "in WT controls could yield false positives if there's any background activity or leaky expression.",
        "Strict Dendra2<-/-> WT removes that risk entirely.",
        "",
        "## EXP replacements made vs. initial draft",
        "Removed: 503, 480 ♀ and 506, 475 ♂  (all mTmG<mTmG/mTmG>; Adipoq-Cre+)",
        "Added:   525, 519 ♀ (young clean)  +  476 ♂ (young clean)  +  347 ♂ + 351/352/353 ♂ (old clean trio)",
        "",
        "## WT pool age distribution (strict Dendra2<-/->)",
        "♀ ≤8mo:  1  (494)",
        "♀ 12–18mo:  9  (119, 177, 161, 165, 330, 333, 341, 378, 202)",
        "♀ >18mo:  1  (00013)",
        "♂ ≤8mo:  0  (after 499 8.3mo is used → none remain in colony)",
        "♂ 8–12mo:  2  (499, 407)",
        "♂ 12–18mo:  7  (172, 317, 320, 322, 324, 325, 381)",
        "♂ >18mo:  4  (68, 82, 00001, 00002)",
        "",
        "Selected WT M = 499, 407, 320★, 322★, 317.  Reserves: 324, 325 (cage 1003059439 siblings).",
        "Selected WT F = 494, 333, 341, 378, 119.  Reserves: 161, 177, 165 (165 used in LPS Ex 2).",
        "",
        "## Cages that will empty after this pull",
        "1003059440 (335, 337) — both pulled",
        "1003121755 (351, 352, 353) — all three pulled, cage empties",
        "1003318024 (503 NOT pulled, kept) — no impact",
        "1003121754 (347 alone) — empties",
        "1003318023 (506 NOT pulled) — no impact",
        "1003225813 (494, 471) — 494 pulled, 471 remains (single-housed female)",
        "",
        "## Reserved / not used (could substitute on request)",
        "EXP backups: 363, 365, 366 ♀ old clean (cages 1003121756/69)",
        "             374 ♂ old clean (cage 1003089166)",
        "             293, 294, 295, 100, 101, 102 ♀ very-old EXP (>14 mo)",
        "WT backups:  324, 325 ♂ strict (cage 1003059439, siblings of 320, 322)",
        "             177, 161 ♀ strict",
    ])

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"[lang] wrote {out_path}")

# ---------- Saved-cohorts JSON ----------

def write_saved_cohorts(out_path: Path):
    payload = {
        "version": 1,
        "generated_at": TODAY,
        "cohorts": [
            {
                "id": "lps_ex2_2026-06-05",
                "name": "LPS Ex 2",
                "owner": "Brian (in-house)",
                "purpose": "In-vivo LPS challenge → peripheral blood Dendra2+ uptake by circulating cells",
                "wt_definition": "option #1 — any mouse without Adipoq-Cre",
                "design": "12 EXP (6 PBS + 6 LPS) + 6 WT gate ctrls (3 PBS + 3 LPS)",
                "exp": [enrich(e, "EXP") for e in LPS_EXP],
                "wt":  [enrich(e, "WT gate ctrl") for e in LPS_WT],
                "matched_pairs": [
                    ["329", "330"], ["445", "446"], ["166", "165"], ["382", "381"],
                ],
                "notes": [
                    "All EXP are mTmG-cassette-clean (no functional mG green-channel background).",
                    "Cages emptying: 1003121676, 1003172013, 1003059365, 1003121780, 1003054428, 1003089166, 1002947895.",
                ],
            },
            {
                "id": "lang_tibial_fracture_2026-06-05",
                "name": "Annemarie Lang — tibial fracture (mito-transfer)",
                "owner": "Annemarie Lang (collaborator, outgoing)",
                "purpose": "Identify cells recruiting to tibial fracture and detect adipocyte-derived mito-Dendra2 transfer",
                "wt_definition": "option #2 — strict: no Adipoq-Cre AND Dendra2<-/->",
                "design": "10 EXP + 10 WT (5F + 5M each)",
                "exp": [enrich(e, "EXP") for e in LANG_EXP],
                "wt":  [enrich(e, "Strict WT") for e in LANG_WT],
                "matched_pairs": [
                    ["335", "320"], ["337", "322"],
                ],
                "notes": [
                    "All EXP mTmG-cassette-clean — mTmG<mTmG/mTmG>;Cre+ mice would express membrane GFP in",
                    "the same green channel as Dendra2, confounding the fracture-site readout.",
                    "WT age skew (7/10 ≥12 mo) reflects strict-Dendra2<-/-> pool constraint.",
                    "351/352/353 = ♂ sibling trio from one litter (built-in biological-replicate cluster).",
                ],
            },
        ],
    }
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with open(out_path, "w") as f:
        json.dump(payload, f, indent=2)
    print(f"[json] wrote {out_path}")

# ---------- Main ----------

if __name__ == "__main__":
    out_dir = HERE
    build_lps_workbook(out_dir / f"Mouse_Cohort_LPS_Ex2_{TODAY}.xlsx")
    build_lang_workbook(out_dir / f"Mouse_Cohort_Lang_TibialFracture_{TODAY}.xlsx")
    write_saved_cohorts(HERE / "data" / "saved_cohorts.json")
    print("done.")
