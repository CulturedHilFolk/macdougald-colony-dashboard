"""
Build a blank rack template the user fills in by hand.
Layout: 7 rows × 10 columns (A-J), matching the physical rack.

Split each rack across TWO landscape pages (cols A-E and cols F-J)
so each cage cell becomes generous writing room (~2.0" × 2.6").

Each cage cell shows:
  - Position label pre-printed (e.g. "1A") in the top-left
  - Pre-printed labels with blank lines:
      Cage #: ____________
      n: ___    Strain: __________
      IDs:
        ☐ ____________
        ☐ ____________
        ☐ ____________
        ☐ ____________
        ☐ ____________
      Notes: __________________

Output: Rack_Template_<DATE>.xlsx
  - Instructions tab
  - Rack 1 A-E, Rack 1 F-J
  - Rack 2 A-E, Rack 2 F-J
  - Rack 3 A-E, Rack 3 F-J
  - Rack 4 A-E, Rack 4 F-J
  - Off-Rack
"""

from pathlib import Path
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DIR = Path("/Users/briandes/Library/CloudStorage/OneDrive-MichiganMedicine/Desktop/MacDougald Lab/Mouse Colony")
TODAY = date.today().isoformat()
ARIAL = "Arial"

ROWS = 7

THIN = Side(border_style="thin", color="BBBBBB")
THICK = Side(border_style="medium", color="333333")
LABEL_FONT = Font(name=ARIAL, size=8, color="888888", italic=True)
POS_FONT = Font(name=ARIAL, size=11, bold=True, color="305496")

# Each cage cell occupies CELL_ROWS spreadsheet rows
CELL_ROWS = 11
# Internal layout (rows offset within a cell):
#   0: position label  +  "Cage #: ____________________"
#   1: "n: ___    Strain: ______________"
#   2: "IDs:"
#   3-7: 5 ID slots (with checkbox glyph)
#   8: "Notes:"
#   9: notes line
#   10: notes line


def add_rack_half(ws, cols_subset, half_label, rack_label_default=""):
    """cols_subset: list of column letters like ['A','B','C','D','E']."""
    n_cols = len(cols_subset)
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.4
    ws.page_margins.bottom = 0.4
    ws.print_options.horizontalCentered = True

    total_cols = 1 + n_cols  # 1 row-label column + cage columns

    # Title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    ws.cell(row=1, column=1,
            value=f"Mouse Colony Walkthrough — {TODAY}    [{half_label}]")
    ws.cell(row=1, column=1).font = Font(name=ARIAL, size=14, bold=True)
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 22

    # Subtitle line for rack/room/walked-by
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_cols)
    ws.cell(row=2, column=1,
            value=f"Rack: {rack_label_default}_______________     "
                  f"Room: __________     Date: {TODAY}     Walked by: _______________")
    ws.cell(row=2, column=1).font = Font(name=ARIAL, size=11)
    ws.cell(row=2, column=1).alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 4

    # Column header (A-E or F-J)
    header_row = 4
    hc = ws.cell(row=header_row, column=1, value="Row")
    hc.font = Font(name=ARIAL, size=11, bold=True, color="FFFFFF")
    hc.fill = PatternFill("solid", start_color="305496")
    hc.alignment = Alignment(horizontal="center", vertical="center")
    hc.border = Border(left=THICK, right=THICK, top=THICK, bottom=THICK)

    for ci, col in enumerate(cols_subset, start=2):
        c = ws.cell(row=header_row, column=ci, value=col)
        c.font = Font(name=ARIAL, size=12, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", start_color="305496")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = Border(left=THICK, right=THICK, top=THICK, bottom=THICK)
    ws.row_dimensions[header_row].height = 20

    # Column widths
    ws.column_dimensions["A"].width = 5
    for ci in range(2, total_cols + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 31

    grid_start = header_row + 1

    # Build cage cells
    for ri in range(ROWS):
        cell_top = grid_start + ri * CELL_ROWS

        # Row label spanning the cell
        ws.merge_cells(start_row=cell_top, start_column=1,
                       end_row=cell_top + CELL_ROWS - 1, end_column=1)
        rl = ws.cell(row=cell_top, column=1, value=str(ri + 1))
        rl.font = Font(name=ARIAL, size=18, bold=True, color="305496")
        rl.alignment = Alignment(horizontal="center", vertical="center")
        rl.fill = PatternFill("solid", start_color="D9E1F2")
        rl.border = Border(left=THICK, right=THICK, top=THICK, bottom=THICK)

        for ci, col_letter in enumerate(cols_subset, start=2):
            pos = f"{ri + 1}{col_letter}"

            # Cell body — write into specific rows, then border the perimeter
            # Row 0: "1A    Cage #: ____________"
            top_cell = ws.cell(row=cell_top, column=ci,
                               value=f"{pos}    Cage #: ")
            top_cell.font = POS_FONT
            top_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)

            # Row 1: "n: ___    Strain: ______"
            ws.cell(row=cell_top + 1, column=ci,
                    value="n: ____      Strain: ____________________").font = LABEL_FONT
            ws.cell(row=cell_top + 1, column=ci).alignment = \
                Alignment(horizontal="left", vertical="center", indent=1)

            # Row 2: "IDs:"
            ws.cell(row=cell_top + 2, column=ci, value="IDs:").font = \
                Font(name=ARIAL, size=9, color="555555", bold=True)
            ws.cell(row=cell_top + 2, column=ci).alignment = \
                Alignment(horizontal="left", vertical="center", indent=1)

            # Rows 3-7: 5 ID slots — bullet glyph
            for slot in range(5):
                r = cell_top + 3 + slot
                ws.cell(row=r, column=ci, value="  •  ").font = \
                    Font(name=ARIAL, size=10, color="AAAAAA")
                ws.cell(row=r, column=ci).alignment = \
                    Alignment(horizontal="left", vertical="center", indent=1)

            # Row 8: "Notes:"
            ws.cell(row=cell_top + 8, column=ci, value="Notes:").font = \
                Font(name=ARIAL, size=9, color="555555", bold=True)
            ws.cell(row=cell_top + 8, column=ci).alignment = \
                Alignment(horizontal="left", vertical="center", indent=1)

            # Borders: thick perimeter, thin internal between rows
            for r in range(cell_top, cell_top + CELL_ROWS):
                cur = ws.cell(row=r, column=ci)
                left = THICK
                right = THICK
                top = THICK if r == cell_top else None
                bottom = THICK if r == cell_top + CELL_ROWS - 1 else None
                # Light separator between header line (row 0) and the rest
                cur.border = Border(left=left, right=right, top=top, bottom=bottom)

        # Row heights inside each cell
        ws.row_dimensions[cell_top].height = 22       # Cage # line
        ws.row_dimensions[cell_top + 1].height = 16   # n / Strain
        ws.row_dimensions[cell_top + 2].height = 14   # "IDs:" header
        for slot in range(5):
            ws.row_dimensions[cell_top + 3 + slot].height = 18
        ws.row_dimensions[cell_top + 8].height = 14   # "Notes:" header
        ws.row_dimensions[cell_top + 9].height = 18
        ws.row_dimensions[cell_top + 10].height = 18

    # Footer
    last_row = grid_start + ROWS * CELL_ROWS
    ws.merge_cells(start_row=last_row, start_column=1,
                   end_row=last_row, end_column=total_cols)
    ws.cell(row=last_row, column=1,
            value=("Page totals — cages: ____   |   mice: ____   |   "
                   f"checked: ☐   |   {half_label}"))
    ws.cell(row=last_row, column=1).font = Font(name=ARIAL, size=10, italic=True)
    ws.cell(row=last_row, column=1).alignment = Alignment(horizontal="center")
    ws.row_dimensions[last_row].height = 18

    ws.print_area = f"A1:{get_column_letter(total_cols)}{last_row}"


def make_offrack_sheet(ws):
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    ws.merge_cells("A1:F1")
    ws["A1"] = f"Off-Rack Cages — {TODAY}"
    ws["A1"].font = Font(name=ARIAL, size=14, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 22

    ws.merge_cells("A2:F2")
    ws["A2"] = ("Use this sheet for any cage that's not in a rack slot "
                "(cart, holding shelf, transferred, etc.)")
    ws["A2"].font = Font(name=ARIAL, size=10, italic=True, color="555555")
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 16

    headers = ["#", "Cage #", "n Mice", "Mouse IDs", "Strain", "Notes / Location"]
    widths = [4, 14, 7, 38, 14, 30]
    HEADER_FILL = PatternFill("solid", start_color="305496")
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    for i, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=i, value=h)
        c.font = Font(name=ARIAL, size=10, bold=True, color="FFFFFF")
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDER
        ws.column_dimensions[get_column_letter(i)].width = widths[i - 1]
    ws.row_dimensions[4].height = 22

    for r in range(5, 60):
        ws.cell(row=r, column=1, value=r - 4).font = \
            Font(name=ARIAL, size=10, color="888888")
        ws.cell(row=r, column=1).alignment = Alignment(horizontal="center", vertical="center")
        for cc in range(1, 7):
            ws.cell(row=r, column=cc).border = BORDER
            ws.cell(row=r, column=cc).font = Font(name=ARIAL, size=10)
            ws.cell(row=r, column=cc).alignment = Alignment(vertical="center", wrap_text=True)
        ws.row_dimensions[r].height = 28


def make_instructions(ws):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 110
    lines = [
        ("Mouse Colony Walkthrough — Blank Rack Template", "title"),
        (f"{TODAY}", "italic"),
        ("", ""),
        ("HOW TO USE", "header"),
        ("", ""),
        ("Each physical rack is split across TWO printed pages: columns A-E and columns F-J.", ""),
        ("This gives roughly 2 inches of writing room per cage cell.", ""),
        ("", ""),
        ("1. Print the rack tabs you need. Each tab fits one landscape page.", ""),
        ("", ""),
        ("2. At the top of each printed page, write:", ""),
        ("     Rack:        the physical rack tag (e.g., '7A')", ""),
        ("     Room:        room name", ""),
        ("     Walked by:   your initials", ""),
        ("", ""),
        ("3. For each occupied slot, fill in:", ""),
        ("     Cage #       — number on the cage card", ""),
        ("     n            — count of mice in the cage", ""),
        ("     Strain       — short strain label (e.g., AdipoGlo, Dendra2)", ""),
        ("     IDs          — list every mouse ID inside (5 lines provided per cage)", ""),
        ("     Notes        — anything off (dead mouse, fight wounds, sick, sex unknown, etc.)", ""),
        ("", ""),
        ("4. Leave empty slots blank.", ""),
        ("", ""),
        ("5. Use the 'Off-Rack' tab for any cage not in a rack slot (cart, holding bin, etc.).", ""),
        ("", ""),
        ("6. When done — scan/photograph the sheets, send them back to me. I will:", ""),
        ("     • Digitize the data", ""),
        ("     • Reconcile against the colony software (flag missing / extra / sex-mismatch)", ""),
        ("     • Build a Transnetyx-compatible Excel for genotyping submission", ""),
        ("     • Build the company-quality colony dashboard for the team", ""),
        ("", ""),
        ("FILE STRUCTURE", "header"),
        ("   • Rack 1 A-E / Rack 1 F-J — one rack across two landscape pages", ""),
        ("   • Rack 2, 3, 4 — same format, one per physical rack", ""),
        ("   • Off-Rack — flat list for cages not in a rack slot", ""),
        ("", ""),
        ("If you need more racks, right-click any tab → Move or Copy → check 'Create a copy'.", ""),
    ]

    for i, (txt, kind) in enumerate(lines, 1):
        c = ws.cell(row=i, column=1, value=txt)
        if kind == "title":
            c.font = Font(name=ARIAL, size=16, bold=True)
        elif kind == "italic":
            c.font = Font(name=ARIAL, size=11, italic=True, color="555555")
        elif kind == "header":
            c.font = Font(name=ARIAL, size=12, bold=True, color="305496")
        else:
            c.font = Font(name=ARIAL, size=10)


# ---------- Build workbook ----------
wb = Workbook()

ws_inst = wb.active
ws_inst.title = "Instructions"
make_instructions(ws_inst)

# 4 racks × 2 halves
for rack_n in range(1, 5):
    ws_a = wb.create_sheet(f"Rack {rack_n} A-E")
    add_rack_half(ws_a, list("ABCDE"), half_label=f"Rack {rack_n} — columns A-E")

    ws_b = wb.create_sheet(f"Rack {rack_n} F-J")
    add_rack_half(ws_b, list("FGHIJ"), half_label=f"Rack {rack_n} — columns F-J")

ws_off = wb.create_sheet("Off-Rack")
make_offrack_sheet(ws_off)

out = DIR / f"Rack_Template_{TODAY}.xlsx"
wb.save(out)
print(f"Saved: {out}")
